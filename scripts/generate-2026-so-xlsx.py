import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("scripts/2026-so-review.json") as f:
    data = json.load(f)

wb = Workbook()

# ── Sheet 1: Item Frequency ──
ws1 = wb.active
ws1.title = "Item Frequency"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Arial", size=10)
num_font = Font(name="Arial", size=10)
border = Border(
    bottom=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
)
alt_fill = PatternFill("solid", fgColor="F2F2F2")

headers1 = ["Rank", "Item Name", "SKU", "Times Used (SOs)", "Total Quantity", "Unique SOs", "Equipment?"]
widths1 = [6, 55, 25, 16, 16, 12, 12]

for col, (header, width) in enumerate(zip(headers1, widths1), 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.column_dimensions[get_column_letter(col)].width = width

ws1.row_dimensions[1].height = 22
ws1.auto_filter.ref = f"A1:G{len(data['itemFrequency']) + 1}"

for i, item in enumerate(data["itemFrequency"]):
    row = i + 2
    fill = alt_fill if i % 2 == 1 else None
    vals = [
        i + 1,
        item["name"],
        item["sku"] or "",
        item["times_used"],
        item["total_qty"],
        item["unique_sos"],
        "Yes" if item["is_equipment"] else "No",
    ]
    for col, val in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.border = border
        if fill:
            cell.fill = fill
        if col in (1, 4, 5, 6):
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "#,##0"
        if col == 7:
            cell.alignment = Alignment(horizontal="center")

# ── Sheet 2: Sales Orders ──
ws2 = wb.create_sheet("Sales Orders")

headers2 = ["SO Number", "Reference", "Customer", "Date", "Status", "Location", "Total ($)", "Item Count"]
widths2 = [14, 35, 30, 12, 12, 25, 14, 12]

for col, (header, width) in enumerate(zip(headers2, widths2), 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.column_dimensions[get_column_letter(col)].width = width

ws2.row_dimensions[1].height = 22
ws2.auto_filter.ref = f"A1:H{len(data['salesOrders']) + 1}"

for i, so in enumerate(data["salesOrders"]):
    row = i + 2
    fill = alt_fill if i % 2 == 1 else None
    vals = [
        so["so_number"],
        so["reference"],
        so["customer"],
        so["date"],
        so["status"],
        so["location"],
        so["total"],
        so["item_count"],
    ]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.border = border
        if fill:
            cell.fill = fill
        if col == 7:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        if col == 8:
            cell.alignment = Alignment(horizontal="center")

# ── Sheet 3: All Line Items (flat) ──
ws3 = wb.create_sheet("All Line Items")

headers3 = ["SO Number", "Reference", "Customer", "Date", "Location", "Item Name", "SKU", "Quantity", "Rate ($)", "Amount ($)"]
widths3 = [14, 35, 25, 12, 22, 50, 22, 10, 12, 14]

for col, (header, width) in enumerate(zip(headers3, widths3), 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.column_dimensions[get_column_letter(col)].width = width

ws3.row_dimensions[1].height = 22

row_num = 2
for so in data["salesOrders"]:
    for item in so["items"]:
        fill = alt_fill if (row_num - 2) % 2 == 1 else None
        vals = [
            so["so_number"],
            so["reference"],
            so["customer"],
            so["date"],
            so["location"],
            item["name"],
            item["sku"] or "",
            item["qty"],
            item["rate"],
            item["amount"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col == 8:
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "#,##0"
            if col in (9, 10):
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

ws3.auto_filter.ref = f"A1:J{row_num - 1}"

# ── Sheet 4: Summary ──
ws4 = wb.create_sheet("Summary")
ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 20

summary_data = [
    ("2026 Sales Order Review", ""),
    ("", ""),
    ("Total Sales Orders", data["summary"]["totalSOs"]),
    ("Total Line Items", data["summary"]["totalLineItems"]),
    ("Unique Products", data["summary"]["uniqueItems"]),
    ("Date Range", data["summary"]["dateRange"]),
]

title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
label_font = Font(name="Arial", bold=True, size=11)
value_font = Font(name="Arial", size=11)

for i, (label, val) in enumerate(summary_data):
    row = i + 1
    cell_a = ws4.cell(row=row, column=1, value=label)
    cell_b = ws4.cell(row=row, column=2, value=val)
    if i == 0:
        cell_a.font = title_font
    elif i >= 2:
        cell_a.font = label_font
        cell_b.font = value_font
        if isinstance(val, (int, float)):
            cell_b.number_format = "#,##0"

# Move Summary to first position
wb.move_sheet(ws4, offset=-3)

out = "scripts/2026-SO-Review.xlsx"
wb.save(out)
print(f"Saved to {out}")
