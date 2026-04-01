import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("scripts/2026-so-inventory-comparison.json") as f:
    data = json.load(f)

wb = Workbook()

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Arial", size=10)
alt_fill = PatternFill("solid", fgColor="F2F2F2")
border = Border(bottom=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"))

green_fill = PatternFill("solid", fgColor="E2EFDA")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
red_fill = PatternFill("solid", fgColor="FCE4EC")

# ── Sheet 1: Summary ──
ws1 = wb.active
ws1.title = "Summary"
ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 15

s = data["summary"]
rows = [
    ("2026 SO vs Internal Inventory", ""),
    ("", ""),
    ("SO Items on 2+ SOs", s["soItemsOn2PlusSOs"]),
    ("Matched to InternalProduct", s["matched"]),
    ("NOT Found in Inventory", s["notFound"]),
    ("", ""),
    ("Total Active InternalProducts", s["totalInternalProducts"]),
    ("InternalProducts NOT on any 2026 SO", s["unusedInternalProducts"]),
    ("", ""),
    ("Match Rate", s.get("matchRate", f"{s['matched'] / s['soItemsOn2PlusSOs'] * 100:.1f}%")),
    ("Gap (missing from inventory)", f"{s['notFound'] / s['soItemsOn2PlusSOs'] * 100:.1f}%"),
]

title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
label_font = Font(name="Arial", bold=True, size=11)
value_font = Font(name="Arial", size=11)

for i, (label, val) in enumerate(rows):
    r = i + 1
    c1 = ws1.cell(row=r, column=1, value=label)
    c2 = ws1.cell(row=r, column=2, value=val)
    if i == 0:
        c1.font = title_font
    elif label:
        c1.font = label_font
        c2.font = value_font
        if isinstance(val, int):
            c2.number_format = "#,##0"

# ── Sheet 2: Full Comparison ──
ws2 = wb.create_sheet("SO Items vs Inventory")

headers = [
    "SO Item Name", "SO SKU", "Times Used", "Total Qty", "Unique SOs",
    "Match Status", "Match Method",
    "IP Category", "IP Brand", "IP Model", "IP Name", "IP SKU",
    "Zoho Linked", "HubSpot Linked", "Zuper Linked",
    "Unit Cost", "Sell Price"
]
widths = [45, 22, 11, 11, 11, 12, 22, 14, 16, 20, 30, 18, 12, 14, 12, 11, 11]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.row_dimensions[1].height = 22

for i, item in enumerate(data["comparison"]):
    row = i + 2
    status = item["match_status"]

    if status == "MATCHED":
        row_fill = green_fill
    elif status == "PARTIAL":
        row_fill = yellow_fill
    else:
        row_fill = red_fill

    vals = [
        item["so_item_name"], item["so_sku"], item["times_used"],
        item["total_qty"], item["unique_sos"],
        status, item["match_method"],
        item["ip_category"], item["ip_brand"], item["ip_model"],
        item["ip_name"], item["ip_sku"],
        "Yes" if item["ip_zoho_linked"] else ("" if not item["ip_id"] else "No"),
        "Yes" if item["ip_hubspot_linked"] else ("" if not item["ip_id"] else "No"),
        "Yes" if item["ip_zuper_linked"] else ("" if not item["ip_id"] else "No"),
        item["ip_unit_cost"], item["ip_sell_price"],
    ]

    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = border
        if col in (3, 4, 5):
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "#,##0"
        if col == 6:
            cell.alignment = Alignment(horizontal="center")
        if col in (13, 14, 15):
            cell.alignment = Alignment(horizontal="center")
        if col in (16, 17) and val is not None:
            cell.number_format = "$#,##0.00"

ws2.auto_filter.ref = f"A1:Q{len(data['comparison']) + 1}"

# ── Sheet 3: Missing from Inventory (NOT_FOUND only) ──
ws3 = wb.create_sheet("Missing from Inventory")

missing_headers = ["SO Item Name", "SO SKU", "Times Used", "Total Qty", "Unique SOs"]
missing_widths = [55, 28, 12, 12, 12]

for col, (h, w) in enumerate(zip(missing_headers, missing_widths), 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="C0392B")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.column_dimensions[get_column_letter(col)].width = w

ws3.row_dimensions[1].height = 22
missing = [c for c in data["comparison"] if c["match_status"] == "NOT_FOUND"]

for i, item in enumerate(missing):
    row = i + 2
    fill = alt_fill if i % 2 == 1 else None
    vals = [item["so_item_name"], item["so_sku"], item["times_used"], item["total_qty"], item["unique_sos"]]
    for col, val in enumerate(vals, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.border = border
        if fill:
            cell.fill = fill
        if col in (3, 4, 5):
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "#,##0"

ws3.auto_filter.ref = f"A1:E{len(missing) + 1}"

# ── Sheet 4: Unused InternalProducts ──
ws4 = wb.create_sheet("Unused InternalProducts")

unused_headers = ["Category", "Brand", "Model", "Name", "SKU", "Zoho", "HubSpot", "Zuper"]
unused_widths = [14, 20, 22, 35, 20, 8, 10, 8]

for col, (h, w) in enumerate(zip(unused_headers, unused_widths), 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="7F8C8D")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws4.column_dimensions[get_column_letter(col)].width = w

ws4.row_dimensions[1].height = 22

for i, p in enumerate(data["unusedProducts"]):
    row = i + 2
    fill = alt_fill if i % 2 == 1 else None
    vals = [
        p["category"], p["brand"], p["model"], p["name"], p["sku"],
        "Yes" if p["zoho_linked"] else "No",
        "Yes" if p["hubspot_linked"] else "No",
        "Yes" if p["zuper_linked"] else "No",
    ]
    for col, val in enumerate(vals, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.border = border
        if fill:
            cell.fill = fill
        if col in (6, 7, 8):
            cell.alignment = Alignment(horizontal="center")

ws4.auto_filter.ref = f"A1:H{len(data['unusedProducts']) + 1}"

out = "scripts/2026-SO-Inventory-Comparison.xlsx"
wb.save(out)
print(f"Saved to {out}")
