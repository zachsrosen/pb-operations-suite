#!/usr/bin/env python3
"""Generate 4x10 Crew Rotation Analysis workbook from JSON data."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Load data ──
with open("scripts/4x10-analysis-data.json") as f:
    data = json.load(f)

meta = data["metadata"]
current = data["currentState"]
proposed = data["proposedModel"]
installs = data["installs"]

# ── Shared styles (matching reference script conventions) ──
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
subtitle_font = Font(name="Arial", size=11, color="666666")
section_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="2F5496")
accent_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
accent_fill = PatternFill("solid", fgColor="2F5496")
totals_fill = PatternFill("solid", fgColor="2F5496")
totals_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
border = Border(
    bottom=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
)
alt_fill = PatternFill("solid", fgColor="F2F2F2")
green_fill = PatternFill("solid", fgColor="C6EFCE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
red_fill = PatternFill("solid", fgColor="FFC7CE")
green_on = PatternFill("solid", fgColor="A9D18E")
yellow_tbd = PatternFill("solid", fgColor="FFD966")
best_fill = PatternFill("solid", fgColor="C6EFCE")

FIT_FILLS = {
    "fits_in_block": green_fill,
    "fits_with_pause": yellow_fill,
    "needs_handoff": red_fill,
}
FIT_LABELS = {
    "fits_in_block": "Fits in Block",
    "fits_with_pause": "Fits with Pause",
    "needs_handoff": "Needs Handoff",
}


def write_header_row(ws, row, headers, widths):
    """Write a styled header row and set column widths."""
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 22


def write_data_cell(ws, row, col, value, fmt=None, align=None, fill=None, font=None):
    """Write a single data cell with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or data_font
    cell.border = border
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = Alignment(horizontal=align)
    if fill:
        cell.fill = fill
    return cell


def write_section_title(ws, row, title, num_cols):
    """Write a merged section title row."""
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=num_cols
    )
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26
    # Fill merged area
    for c in range(2, num_cols + 1):
        ws.cell(row=row, column=c).fill = section_fill


# ── Create workbook ──
wb = Workbook()


# ============================================================
# TAB 1 — Current State (5x8)
# ============================================================
ws1 = wb.active
ws1.title = "Current State (5x8)"

# Section A: Location Summary Table
headers_a = [
    "Location", "Installs", "Avg Duration (days)", "Median Duration",
    "Total Revenue", "Avg Revenue", "Utilization"
]
widths_a = [22, 12, 18, 16, 18, 16, 14]
write_header_row(ws1, 1, headers_a, widths_a)
ws1.freeze_panes = "A2"

loc_summaries = current["locationSummaries"]
for i, loc in enumerate(loc_summaries):
    row = i + 2
    fill = alt_fill if i % 2 == 1 else None
    vals = [
        (loc["location"], None, "left"),
        (loc["totalInstalls"], "#,##0", "center"),
        (loc["avgCrewDays"], "0.0", "center"),
        (loc["medianCrewDays"], "0.0", "center"),
        (loc["totalRevenue"], '"$"#,##0', "right"),
        (loc["avgRevenue"], '"$"#,##0', "right"),
        (loc["utilization"], "0.0%", "center"),
    ]
    for col, (val, fmt, align) in enumerate(vals, 1):
        # Utilization is stored as e.g. 199.6 meaning 199.6% — convert to fraction
        if col == 7:
            val = val / 100.0
        write_data_cell(ws1, row, col, val, fmt=fmt, align=align, fill=fill)

# Totals row
totals_row = len(loc_summaries) + 2
total_installs = sum(l["totalInstalls"] for l in loc_summaries)
total_revenue = sum(l["totalRevenue"] for l in loc_summaries)
avg_crew = current["overall"]["avgCrewDays"]
median_crew = current["overall"]["medianCrewDays"]
avg_rev = total_revenue / total_installs if total_installs else 0
total_util = sum(l["utilization"] for l in loc_summaries) / len(loc_summaries) / 100.0

totals_vals = [
    ("TOTAL", None, "left"),
    (total_installs, "#,##0", "center"),
    (avg_crew, "0.0", "center"),
    (median_crew, "0.0", "center"),
    (total_revenue, '"$"#,##0', "right"),
    (avg_rev, '"$"#,##0', "right"),
    (total_util, "0.0%", "center"),
]
for col, (val, fmt, align) in enumerate(totals_vals, 1):
    write_data_cell(ws1, totals_row, col, val, fmt=fmt, align=align,
                    fill=totals_fill, font=totals_font)

# Section B: Day-of-Week Distribution
sec_b_start = totals_row + 3
write_section_title(ws1, sec_b_start, "INSTALL START DAY DISTRIBUTION", 7)

dow_headers = ["Location", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
dow_row = sec_b_start + 1
for col, h in enumerate(dow_headers, 1):
    cell = ws1.cell(row=dow_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for i, loc in enumerate(loc_summaries):
    row = dow_row + 1 + i
    fill = alt_fill if i % 2 == 1 else None
    dist = loc["dayOfWeekDistribution"]
    write_data_cell(ws1, row, 1, loc["location"], align="left", fill=fill)
    for j, day in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], 2):
        write_data_cell(ws1, row, j, dist.get(day, 0), fmt="#,##0", align="center", fill=fill)

# Section C: Monthly Breakdown
sec_c_start = dow_row + 1 + len(loc_summaries) + 2
write_section_title(ws1, sec_c_start, "MONTHLY BREAKDOWN", 5)

monthly_headers = ["Month", "Location", "Installs", "Crew-Days Used", "Revenue"]
monthly_widths = [14, 22, 12, 16, 18]
monthly_hdr_row = sec_c_start + 1
for col, (h, w) in enumerate(zip(monthly_headers, monthly_widths), 1):
    cell = ws1.cell(row=monthly_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Collect all months across locations
all_months = set()
for loc in loc_summaries:
    all_months.update(loc["monthlyBreakdown"].keys())
sorted_months = sorted(all_months)

r = monthly_hdr_row + 1
row_idx = 0
for month in sorted_months:
    for loc in loc_summaries:
        mb = loc["monthlyBreakdown"].get(month)
        if not mb or mb["count"] == 0:
            continue
        fill = alt_fill if row_idx % 2 == 1 else None
        write_data_cell(ws1, r, 1, month, align="center", fill=fill)
        write_data_cell(ws1, r, 2, loc["location"], align="left", fill=fill)
        write_data_cell(ws1, r, 3, mb["count"], fmt="#,##0", align="center", fill=fill)
        write_data_cell(ws1, r, 4, mb["crewDays"], fmt="#,##0", align="center", fill=fill)
        write_data_cell(ws1, r, 5, mb["revenue"], fmt='"$"#,##0', align="right", fill=fill)
        r += 1
        row_idx += 1

# Section D: Weekly Capacity Headline
sec_d_start = r + 2
write_section_title(ws1, sec_d_start, "WEEKLY CAPACITY", 2)

cap_data = [
    ("Crew-days/week", current["overall"]["totalCrewDaysPerWeek"]),
    ("Crew-hours/week", current["overall"]["totalCrewHoursPerWeek"]),
]
for i, (label, val) in enumerate(cap_data):
    row = sec_d_start + 1 + i
    write_data_cell(ws1, row, 1, label, font=bold_font, align="left")
    write_data_cell(ws1, row, 2, val, fmt="#,##0", align="center", font=bold_font)


# ============================================================
# TAB 2 — Proposed Model (4x10)
# ============================================================
ws2 = wb.create_sheet("Proposed Model (4x10)")

# Section 1: Coverage Calendar
cal_headers = ["Location", "Crew", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
cal_widths = [22, 14, 10, 10, 10, 10, 10, 10]
write_header_row(ws2, 1, cal_headers, cal_widths)

# Define crew schedule: (location, crew, mon, tue, wed, thu, fri, sat)
# Group A = Mon-Thu, Group B = Wed-Sat
crew_schedule = [
    ("Westminster", "Crew 1 (A)", ["ON", "ON", "ON", "ON", "", ""]),
    ("Westminster", "Crew 2 (B)", ["", "", "ON", "ON", "ON", "ON"]),
    ("Centennial", "Crew 1 (A)", ["ON", "ON", "ON", "ON", "", ""]),
    ("Centennial", "Crew 2 (B)", ["", "", "ON", "ON", "ON", "ON"]),
    ("Colorado Springs", "Crew 1 (TBD)", ["?", "?", "?", "?", "?", "?"]),
]

for i, (loc, crew, days) in enumerate(crew_schedule):
    row = i + 2
    fill = alt_fill if i % 2 == 1 else None
    write_data_cell(ws2, row, 1, loc, align="left", fill=fill)
    write_data_cell(ws2, row, 2, crew, align="center", fill=fill)
    for j, day_val in enumerate(days, 3):
        if day_val == "ON":
            write_data_cell(ws2, row, j, "ON", align="center", fill=green_on,
                            font=Font(name="Arial", bold=True, size=10, color="375623"))
        elif day_val == "?":
            write_data_cell(ws2, row, j, "?", align="center", fill=yellow_tbd,
                            font=Font(name="Arial", bold=True, size=10))
        else:
            write_data_cell(ws2, row, j, "", align="center", fill=fill)

# Overlap note
overlap_row = 2 + len(crew_schedule) + 1
ws2.merge_cells(start_row=overlap_row, start_column=1, end_row=overlap_row, end_column=8)
note_cell = ws2.cell(row=overlap_row, column=1,
                     value="Note: Wed-Thu are overlap days where both Group A and Group B are active (handoff buffer)")
note_cell.font = Font(name="Arial", italic=True, size=10, color="666666")

# Section 2: Install Fit Analysis (Pause Allowed)
sec2_start = overlap_row + 2
write_section_title(ws2, sec2_start, "INSTALL FIT ANALYSIS \u2014 PAUSE ALLOWED", 3)

fit_pa = proposed["fitDistribution"]["pauseAllowed"]
fit_headers = ["Classification", "Count", "Percentage"]
fit_hdr_row = sec2_start + 1
for col, h in enumerate(fit_headers, 1):
    cell = ws2.cell(row=fit_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

total_pa = fit_pa["fitsInBlock"] + fit_pa["fitsWithPause"] + fit_pa["needsHandoff"]
fit_pa_rows = [
    ("Fits in Block", fit_pa["fitsInBlock"], green_fill),
    ("Fits with Pause", fit_pa["fitsWithPause"], yellow_fill),
    ("Needs Handoff", fit_pa["needsHandoff"], red_fill),
]
for i, (label, count, cfill) in enumerate(fit_pa_rows):
    row = fit_hdr_row + 1 + i
    write_data_cell(ws2, row, 1, label, align="left", fill=cfill)
    write_data_cell(ws2, row, 2, count, fmt="#,##0", align="center", fill=cfill)
    write_data_cell(ws2, row, 3, count / total_pa if total_pa else 0, fmt="0.0%", align="center", fill=cfill)

# Total row
pa_total_row = fit_hdr_row + 1 + len(fit_pa_rows)
write_data_cell(ws2, pa_total_row, 1, "TOTAL", align="left", fill=totals_fill, font=totals_font)
write_data_cell(ws2, pa_total_row, 2, total_pa, fmt="#,##0", align="center", fill=totals_fill, font=totals_font)
write_data_cell(ws2, pa_total_row, 3, 1.0, fmt="0.0%", align="center", fill=totals_fill, font=totals_font)

# Section 3: Install Fit Analysis (No Pause)
sec3_start = pa_total_row + 2
write_section_title(ws2, sec3_start, "INSTALL FIT ANALYSIS \u2014 NO PAUSE", 3)

fit_np = proposed["fitDistribution"]["noPause"]
fit_np_hdr_row = sec3_start + 1
for col, h in enumerate(fit_headers, 1):
    cell = ws2.cell(row=fit_np_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

total_np = fit_np["fitsInBlock"] + fit_np["fitsWithPause"] + fit_np["needsHandoff"]
fit_np_rows = [
    ("Fits in Block", fit_np["fitsInBlock"], green_fill),
    ("Needs Handoff", fit_np["needsHandoff"], red_fill),
]
for i, (label, count, cfill) in enumerate(fit_np_rows):
    row = fit_np_hdr_row + 1 + i
    write_data_cell(ws2, row, 1, label, align="left", fill=cfill)
    write_data_cell(ws2, row, 2, count, fmt="#,##0", align="center", fill=cfill)
    write_data_cell(ws2, row, 3, count / total_np if total_np else 0, fmt="0.0%", align="center", fill=cfill)

np_total_row = fit_np_hdr_row + 1 + len(fit_np_rows)
write_data_cell(ws2, np_total_row, 1, "TOTAL", align="left", fill=totals_fill, font=totals_font)
write_data_cell(ws2, np_total_row, 2, total_np, fmt="#,##0", align="center", fill=totals_fill, font=totals_font)
write_data_cell(ws2, np_total_row, 3, 1.0, fmt="0.0%", align="center", fill=totals_fill, font=totals_font)

# Section 4: Calendar Span Comparison
sec4_start = np_total_row + 2
write_section_title(ws2, sec4_start, "AVERAGE CALENDAR SPAN COMPARISON", 4)

span_headers = ["Metric", "Current (5x8)", "Proposed (4x10)", "Change"]
span_hdr_row = sec4_start + 1
span_widths = [28, 16, 16, 12]
for col, (h, w) in enumerate(zip(span_headers, span_widths), 1):
    cell = ws2.cell(row=span_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    ws2.column_dimensions[get_column_letter(col)].width = max(
        ws2.column_dimensions[get_column_letter(col)].width or 0, w
    )

# Compute averages
current_avg_elapsed = sum(inst["elapsedCalendarDays"] for inst in installs) / len(installs) if installs else 0
proposed_avg_elapsed = sum(
    inst["simPauseAllowed"]["calendarDaysToComplete"] for inst in installs
) / len(installs) if installs else 0
current_avg_crew = sum(inst["crewDaysRequired"] for inst in installs) / len(installs) if installs else 0

span_data = [
    ("Avg elapsed calendar days", current_avg_elapsed, proposed_avg_elapsed),
    ("Avg crew-days required", current_avg_crew, current_avg_crew),  # same in both
]
for i, (label, cur, prop) in enumerate(span_data):
    row = span_hdr_row + 1 + i
    change = prop - cur
    change_fill = green_fill if change < 0 else (red_fill if change > 0 else None)
    write_data_cell(ws2, row, 1, label, align="left")
    write_data_cell(ws2, row, 2, cur, fmt="0.0", align="center")
    write_data_cell(ws2, row, 3, prop, fmt="0.0", align="center")
    write_data_cell(ws2, row, 4, change, fmt="+0.0;-0.0;0.0", align="center", fill=change_fill)

# Section 5: Capacity Comparison
sec5_start = span_hdr_row + 1 + len(span_data) + 1
write_section_title(ws2, sec5_start, "CAPACITY COMPARISON", 4)

cap_headers = ["Metric", "Current (5x8)", "Proposed (4x10)", "Change"]
cap_hdr_row = sec5_start + 1
for col, h in enumerate(cap_headers, 1):
    cell = ws2.cell(row=cap_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

cap_data_rows = [
    ("Crew-days/week", current["overall"]["totalCrewDaysPerWeek"],
     proposed["overall"]["totalCrewDaysPerWeek"]),
    ("Crew-hours/week", current["overall"]["totalCrewHoursPerWeek"],
     proposed["overall"]["totalCrewHoursPerWeek"]),
    ("Operating days/week", 5, proposed["overall"]["operatingDays"]),
]
for i, (label, cur, prop) in enumerate(cap_data_rows):
    row = cap_hdr_row + 1 + i
    change = prop - cur
    if label == "Operating days/week":
        change_fill = green_fill if change > 0 else (red_fill if change < 0 else None)
    else:
        change_fill = green_fill if change > 0 else (red_fill if change < 0 else None)
    # For crew-days, fewer is neutral/negative; for operating days, more is positive
    # Actually: fewer crew-days means less capacity (red), more operating days is good (green)
    if label == "Crew-days/week":
        change_fill = red_fill if change < 0 else (green_fill if change > 0 else None)
    elif label == "Crew-hours/week":
        change_fill = green_fill if change > 0 else (red_fill if change < 0 else None)

    write_data_cell(ws2, row, 1, label, align="left")
    write_data_cell(ws2, row, 2, cur, fmt="#,##0", align="center")
    write_data_cell(ws2, row, 3, prop, fmt="#,##0", align="center")
    fmt_str = "+#,##0;-#,##0;0"
    write_data_cell(ws2, row, 4, change, fmt=fmt_str, align="center", fill=change_fill)

# Section 6: Per-Install Detail Table
sec6_start = cap_hdr_row + 1 + len(cap_data_rows) + 1
write_section_title(ws2, sec6_start, "PER-INSTALL DETAIL", 12)

detail_headers = [
    "Project #", "Deal Name", "Location", "Amount", "Schedule Date",
    "Complete Date", "Crew-Days", "Start Day", "Group",
    "Fit (Pause)", "Fit (No Pause)", "Sim Calendar Days"
]
detail_widths = [14, 50, 18, 14, 14, 14, 12, 12, 10, 16, 16, 16]
detail_hdr_row = sec6_start + 1
write_header_row(ws2, detail_hdr_row, detail_headers, detail_widths)

for i, inst in enumerate(installs):
    row = detail_hdr_row + 1 + i
    fill = alt_fill if i % 2 == 1 else None

    fit_pause_class = inst["simPauseAllowed"]["fitClassification"]
    fit_nopause_class = inst["simNoPause"]["fitClassification"]

    write_data_cell(ws2, row, 1, inst["projectNumber"], align="left", fill=fill)
    write_data_cell(ws2, row, 2, inst["dealName"], align="left", fill=fill)
    write_data_cell(ws2, row, 3, inst["location"], align="center", fill=fill)
    write_data_cell(ws2, row, 4, inst["amount"], fmt='"$"#,##0', align="right", fill=fill)
    write_data_cell(ws2, row, 5, inst["installScheduleDate"], align="center", fill=fill)
    write_data_cell(ws2, row, 6, inst["constructionCompleteDate"], align="center", fill=fill)
    write_data_cell(ws2, row, 7, inst["crewDaysRequired"], fmt="#,##0", align="center", fill=fill)
    write_data_cell(ws2, row, 8, inst["startDayName"], align="center", fill=fill)
    write_data_cell(ws2, row, 9, inst["assignedGroup"], align="center", fill=fill)
    write_data_cell(ws2, row, 10, FIT_LABELS.get(fit_pause_class, fit_pause_class),
                    align="center", fill=FIT_FILLS.get(fit_pause_class, fill))
    write_data_cell(ws2, row, 11, FIT_LABELS.get(fit_nopause_class, fit_nopause_class),
                    align="center", fill=FIT_FILLS.get(fit_nopause_class, fill))
    write_data_cell(ws2, row, 12, inst["simPauseAllowed"]["calendarDaysToComplete"],
                    fmt="#,##0", align="center", fill=fill)

# Auto-filter on the detail table
last_detail_row = detail_hdr_row + len(installs)
ws2.auto_filter.ref = f"A{detail_hdr_row}:L{last_detail_row}"


# ============================================================
# TAB 3 — Scenarios
# ============================================================
ws3 = wb.create_sheet("Scenarios")

# Section 1: COSP Group Assignment
write_section_title(ws3, 1, "COSP GROUP ASSIGNMENT COMPARISON", 6)

scen_headers = ["Scenario", "Install Coverage", "Handoff Rate", "Revenue/Week", "Dark Days", "Score"]
scen_widths = [36, 16, 14, 16, 12, 10]
scen_hdr_row = 2
write_header_row(ws3, scen_hdr_row, scen_headers, scen_widths)
ws3.freeze_panes = "A3"

scenarios = proposed["scenarios"]
best_score_idx = max(range(len(scenarios)), key=lambda x: scenarios[x]["weightedScore"])
for i, sc in enumerate(scenarios):
    row = scen_hdr_row + 1 + i
    is_best = (i == best_score_idx)
    fill = best_fill if is_best else (alt_fill if i % 2 == 1 else None)
    write_data_cell(ws3, row, 1, sc["label"], align="left", fill=fill)
    write_data_cell(ws3, row, 2, sc["installCoverage"], fmt="0.0%", align="center", fill=fill)
    write_data_cell(ws3, row, 3, sc["handoffRate"], fmt="0.0%", align="center", fill=fill)
    write_data_cell(ws3, row, 4, sc["revenueCapacity"], fmt='"$"#,##0', align="right", fill=fill)
    write_data_cell(ws3, row, 5, sc["darkDays"], fmt="#,##0", align="center", fill=fill)
    write_data_cell(ws3, row, 6, sc["weightedScore"], fmt="0.0", align="center", fill=fill)

# Section 2: 6th Crew Placement
sec2_scen_start = scen_hdr_row + 1 + len(scenarios) + 2
write_section_title(ws3, sec2_scen_start, "6TH CREW PLACEMENT OPTIONS (12 SCENARIOS)", 6)

sixth_hdr_row = sec2_scen_start + 1
for col, h in enumerate(scen_headers, 1):
    cell = ws3.cell(row=sixth_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

sixth_scenarios = proposed["sixthCrewScenarios"]
best_sixth_idx = max(range(len(sixth_scenarios)), key=lambda x: sixth_scenarios[x]["weightedScore"])
for i, sc in enumerate(sixth_scenarios):
    row = sixth_hdr_row + 1 + i
    is_best = (i == best_sixth_idx)
    fill = best_fill if is_best else (alt_fill if i % 2 == 1 else None)
    write_data_cell(ws3, row, 1, sc["label"], align="left", fill=fill)
    write_data_cell(ws3, row, 2, sc["installCoverage"], fmt="0.0%", align="center", fill=fill)
    write_data_cell(ws3, row, 3, sc["handoffRate"], fmt="0.0%", align="center", fill=fill)
    write_data_cell(ws3, row, 4, sc["revenueCapacity"], fmt='"$"#,##0', align="right", fill=fill)
    write_data_cell(ws3, row, 5, sc["darkDays"], fmt="#,##0", align="center", fill=fill)
    write_data_cell(ws3, row, 6, sc["weightedScore"], fmt="0.0", align="center", fill=fill)

# Section 3: Turnaround Compression
sec3_scen_start = sixth_hdr_row + 1 + len(sixth_scenarios) + 2
write_section_title(ws3, sec3_scen_start, "TURNAROUND COMPRESSION IMPACT", 5)

comp_headers = ["Compression", "Fits Block", "Fits Pause", "Needs Handoff", "Score"]
comp_hdr_row = sec3_scen_start + 1
for col, (h, w) in enumerate(zip(comp_headers, [22, 14, 14, 14, 10]), 1):
    cell = ws3.cell(row=comp_hdr_row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

comp_scenarios = proposed["compressionScenarios"]
for i, cs in enumerate(comp_scenarios):
    row = comp_hdr_row + 1 + i
    fill = alt_fill if i % 2 == 1 else None
    fd = cs["fitDistPause"]
    write_data_cell(ws3, row, 1, f"-{cs['compressionDays']} day", align="center", fill=fill)
    write_data_cell(ws3, row, 2, fd["fitsInBlock"], fmt="#,##0", align="center", fill=fill)
    write_data_cell(ws3, row, 3, fd["fitsWithPause"], fmt="#,##0", align="center", fill=fill)
    write_data_cell(ws3, row, 4, fd["needsHandoff"], fmt="#,##0", align="center", fill=fill)
    write_data_cell(ws3, row, 5, cs["score"]["weightedScore"], fmt="0.0", align="center", fill=fill)


# ============================================================
# TAB 4 — Executive Summary
# ============================================================
ws4 = wb.create_sheet("Executive Summary")
ws4.column_dimensions["A"].width = 32
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 20
ws4.column_dimensions["D"].width = 14

# Section 1: Headline
title_cell = ws4.cell(row=1, column=1, value="4x10 CREW ROTATION \u2014 EXECUTIVE SUMMARY")
title_cell.font = Font(name="Arial", bold=True, size=16, color="2F5496")
ws4.merge_cells("A1:D1")
ws4.row_dimensions[1].height = 30

date_range = f"{meta['dateRange']['from']} to {meta['dateRange']['to']}"
sub_cell = ws4.cell(row=2, column=1,
                    value=f"Analysis of {meta['validInstalls']} completed installs, {date_range}")
sub_cell.font = subtitle_font
ws4.merge_cells("A2:D2")

# Comparison table
comp_table_start = 4
exec_headers = ["Metric", "Current (5x8)", "Proposed (4x10)", "Change"]
for col, h in enumerate(exec_headers, 1):
    cell = ws4.cell(row=comp_table_start, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

exec_rows = [
    ("Crew-hours/week", 200, 200, "0"),
    ("Crew-days/week", 25, 20, "-5"),
    ("Operating days", 5, 6, "+1"),
    ("Weekend", "2-day (Sat-Sun)", "3-day (varies)", ""),
    ("Saturday coverage", "No", "Yes", ""),
]
for i, (metric, cur, prop, change) in enumerate(exec_rows):
    row = comp_table_start + 1 + i
    fill = alt_fill if i % 2 == 1 else None
    write_data_cell(ws4, row, 1, metric, align="left", fill=fill, font=bold_font)
    write_data_cell(ws4, row, 2, cur, align="center", fill=fill)
    write_data_cell(ws4, row, 3, prop, align="center", fill=fill)
    # Color-code change
    change_fill = fill
    if change.startswith("+"):
        change_fill = green_fill
    elif change.startswith("-"):
        change_fill = red_fill
    write_data_cell(ws4, row, 4, change, align="center", fill=change_fill)

# Section 2: Install Fit Headline
fit_sec_start = comp_table_start + 1 + len(exec_rows) + 2
fit_pa_data = proposed["fitDistribution"]["pauseAllowed"]
same_crew_count = fit_pa_data["fitsInBlock"] + fit_pa_data["fitsWithPause"]
same_crew_total = same_crew_count + fit_pa_data["needsHandoff"]
same_crew_pct = same_crew_count / same_crew_total * 100 if same_crew_total else 0

headline_cell = ws4.cell(row=fit_sec_start, column=1,
                         value=f"{same_crew_pct:.0f}% of installs handled by same crew")
headline_cell.font = Font(name="Arial", bold=True, size=14, color="375623")
ws4.merge_cells(start_row=fit_sec_start, start_column=1,
                end_row=fit_sec_start, end_column=4)

detail_cell = ws4.cell(
    row=fit_sec_start + 1, column=1,
    value=(f"{fit_pa_data['fitsInBlock']} installs fit in block, "
           f"{fit_pa_data['fitsWithPause']} with pause, "
           f"{fit_pa_data['needsHandoff']} need handoff")
)
detail_cell.font = subtitle_font
ws4.merge_cells(start_row=fit_sec_start + 1, start_column=1,
                end_row=fit_sec_start + 1, end_column=4)

# Section 3: Recommendation
rec_start = fit_sec_start + 4
write_section_title(ws4, rec_start, "RECOMMENDED CONFIGURATION", 4)

# Find best base scenario
best_base = max(proposed["scenarios"], key=lambda s: s["weightedScore"])
best_sixth = max(proposed["sixthCrewScenarios"], key=lambda s: s["weightedScore"])

ws4.cell(row=rec_start + 1, column=1, value="Best base scenario:").font = bold_font
ws4.cell(row=rec_start + 1, column=2,
         value=f"{best_base['label']} (score: {best_base['weightedScore']:.1f})").font = data_font
ws4.merge_cells(start_row=rec_start + 1, start_column=2,
                end_row=rec_start + 1, end_column=4)

ws4.cell(row=rec_start + 2, column=1, value="Best 6th crew scenario:").font = bold_font
ws4.cell(row=rec_start + 2, column=2,
         value=f"{best_sixth['label']} (score: {best_sixth['weightedScore']:.1f})").font = data_font
ws4.merge_cells(start_row=rec_start + 2, start_column=2,
                end_row=rec_start + 2, end_column=4)

# Determine COSP group from best base scenario label
cosp_group = "A" if "Group A" in best_base["label"] else "B"
ws4.cell(row=rec_start + 3, column=1, value=f"COSP: Group {cosp_group}").font = bold_font

# Section 4: Advantages
adv_start = rec_start + 5
write_section_title(ws4, adv_start, "ADVANTAGES", 4)

advantages = [
    "Same crew-hours, 6-day operating window",
    "3-day weekends improve crew retention",
    "Wed-Thu overlap provides handoff buffer",
    "Saturday coverage captures weekend demand",
]
for i, item in enumerate(advantages):
    row = adv_start + 1 + i
    ws4.cell(row=row, column=1, value=f"\u2022 {item}").font = data_font
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

# Section 5: Risks & Considerations
risk_start = adv_start + 1 + len(advantages) + 1
write_section_title(ws4, risk_start, "RISKS & CONSIDERATIONS", 4)

handoff_count = fit_pa_data["needsHandoff"]
handoff_pct = handoff_count / same_crew_total * 100 if same_crew_total else 0

risks = [
    f"{handoff_count} installs ({handoff_pct:.0f}%) may need handoff between crews",
    "Colorado Springs has 2 dark days per week (1-crew location)",
    "Longer daily hours may reduce end-of-day productivity",
    "Customer communication needed for multi-day install gaps",
]
for i, item in enumerate(risks):
    row = risk_start + 1 + i
    ws4.cell(row=row, column=1, value=f"\u2022 {item}").font = data_font
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


# ── Move Executive Summary to first position ──
wb.move_sheet(ws4, offset=-3)

# ── Save ──
os.makedirs("reports", exist_ok=True)
out_path = "reports/4x10-crew-rotation-analysis.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")
print(f"  Tabs: {[ws.title for ws in wb.worksheets]}")
print(f"  Installs: {len(installs)}")
print(f"  Scenarios: {len(proposed['scenarios'])} base, {len(proposed['sixthCrewScenarios'])} 6th-crew")
