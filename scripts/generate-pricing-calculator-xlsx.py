#!/usr/bin/env python3
"""Generate PB Pricing Calculator — dropdown-driven, simple input, full breakdown."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Styles
BLUE = Font(name="Arial", color="0000FF", size=11)
BLUE_B = Font(name="Arial", color="0000FF", size=11, bold=True)
BLK = Font(name="Arial", color="000000", size=11)
BLK_B = Font(name="Arial", color="000000", size=11, bold=True)
WHT_B = Font(name="Arial", color="FFFFFF", size=11, bold=True)
HINT = Font(name="Arial", color="718096", size=10, italic=True)
HDR = PatternFill("solid", fgColor="2D3748")
SEC = PatternFill("solid", fgColor="4A5568")
YEL = PatternFill("solid", fgColor="FFFF00")
LGRAY = PatternFill("solid", fgColor="F7FAFC")
LGRN = PatternFill("solid", fgColor="F0FFF4")
LORG = PatternFill("solid", fgColor="FFFAF0")
LBLU = PatternFill("solid", fgColor="EBF4FF")
BDR = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
CUR = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%'
FAC = '0.0000000'
NUM = '#,##0'
RATE = '$#,##0.00'

def sc(ws, r, c, v=None, f=None, bg=None, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f: cell.font = f
    if bg: cell.fill = bg
    if nf: cell.number_format = nf
    cell.border = BDR
    return cell

def sec_hdr(ws, r, txt, cols=5):
    for c in range(1, cols+1):
        cell = sc(ws, r, c, txt if c == 1 else None, WHT_B, SEC)
        cell.alignment = Alignment(horizontal="left")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1: Costs & Prices
# ═══════════════════════════════════════════════════════════════════════════════
cp = wb.active
cp.title = "Costs & Prices"
cp.sheet_properties.tabColor = "38A169"
for c, w in enumerate([30, 14, 14, 14, 14, 14], 1):
    cp.column_dimensions[get_column_letter(c)].width = w

# --- Modules ---
r = 1
sec_hdr(cp, r, "MODULES", 6)
r = 2
for c, h in enumerate(["Name", "Cost", "Watts", "DC Qual?", "", ""], 1):
    sc(cp, r, c, h, BLK_B, LGRAY)
modules = [
    ("Hyundai 440W (Black)", 305, 440, "No"),
    ("Silfab 440W (Black/Gold)", 305, 440, "No"),
    ("Silfab 430W (Black/Gold)", 300, 430, "No"),
]
MOD_S = 3
for i, (n, cost, w, dc) in enumerate(modules):
    rr = MOD_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, cost, BLUE, YEL, CUR)
    sc(cp, rr, 3, w, BLUE, YEL, NUM)
    sc(cp, rr, 4, dc, BLK)
MOD_E = MOD_S + len(modules) - 1

# --- Inverters ---
r = MOD_E + 2
sec_hdr(cp, r, "INVERTERS", 6)
r += 1
for c, h in enumerate(["Name", "Cost", "", "", "", ""], 1):
    sc(cp, r, c, h, BLK_B, LGRAY)
inverters = [
    ("Tesla 7.6kW Inverter", 1200),
    ("Tesla 5.0kW Inverter", 1200),
    ("Tesla 3.8kW Inverter", 1200),
    ("Enphase IQ8MC Micro", 160),
    ("Enphase IQ8X Micro", 158),
    ("Enphase IQ8A Micro", 143),
    ("Enphase IQ8M Micro", 150),
]
INV_S = r + 1
for i, (n, cost) in enumerate(inverters):
    rr = INV_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, cost, BLUE, YEL, CUR)
INV_E = INV_S + len(inverters) - 1

# --- Batteries ---
r = INV_E + 2
sec_hdr(cp, r, "BATTERIES", 6)
r += 1
for c, h in enumerate(["Name", "Cost", "Labour", "DC Qual?", "", ""], 1):
    sc(cp, r, c, h, BLK_B, LGRAY)
batteries = [
    ("Tesla Powerwall 3", 7700, 2600, "Yes"),
    ("PW3 Expansion Pack", 5000, 1900, "Yes"),
]
BAT_S = r + 1
for i, (n, cost, lab, dc) in enumerate(batteries):
    rr = BAT_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, cost, BLUE, YEL, CUR)
    sc(cp, rr, 3, lab, BLUE, YEL, CUR)
    sc(cp, rr, 4, dc, BLK)
BAT_E = BAT_S + len(batteries) - 1

# --- Other Equipment ---
r = BAT_E + 2
sec_hdr(cp, r, "OTHER EQUIPMENT", 6)
r += 1
for c, h in enumerate(["Name", "Cost", "Bundled?", "", "", ""], 1):
    sc(cp, r, c, h, BLK_B, LGRAY)
others = [
    ("Tesla Wall Connector (Gen 3)", 600, "No"),
    ("Tesla Backup Switch", 305, "Yes"),
]
OTH_S = r + 1
for i, (n, cost, bundled) in enumerate(others):
    rr = OTH_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, cost, BLUE, YEL, CUR)
    sc(cp, rr, 3, bundled, BLK)
OTH_E = OTH_S + len(others) - 1

# --- Pricing Schemes ---
r = OTH_E + 2
sec_hdr(cp, r, "PRICING SCHEMES", 3)
r += 1
for c, h in enumerate(["Scheme", "Markup %", ""], 1):
    sc(cp, r, c, h, BLK_B, LGRAY)
schemes = [("Base (Colorado)", 0.40), ("Ventura", 0.36), ("Bay Area", 0.50), ("D&R", 0.30), ("Off Grid Homes", 0.65)]
SCH_S = r + 1
for i, (n, pct) in enumerate(schemes):
    rr = SCH_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, pct, BLUE, YEL, PCT)
SCH_E = SCH_S + len(schemes) - 1

# --- Roof Types ---
r = SCH_E + 2
sec_hdr(cp, r, "ROOF TYPES", 3)
r += 1
for c, h in enumerate(["Type", "Fixed ($)", "Per-Watt ($)"], 1):
    sc(cp, r, c, h, BLK_B, LGRAY)
roofs = [("Comp/Asphalt Shingle", 0, 0), ("Flat Membrane", 0, 0.35), ("Tile Concrete", 3500, 0.80), ("Wood Shake", 0, 0.35), ("Metal Corrugated", 0, 0.35)]
RF_S = r + 1
for i, (n, fx, pw) in enumerate(roofs):
    rr = RF_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, fx, BLUE, YEL, CUR)
    sc(cp, rr, 3, pw, BLUE, YEL, RATE)
RF_E = RF_S + len(roofs) - 1

# --- Storey ---
r = RF_E + 2
sec_hdr(cp, r, "STOREY ADDERS", 2)
r += 1
sc(cp, r, 1, "Stories", BLK_B, LGRAY); sc(cp, r, 2, "Per-Watt ($)", BLK_B, LGRAY)
storeys = [("1 Story", 0), ("2 Stories", 0.05), ("3+ Stories", 0.83)]
ST_S = r + 1
for i, (n, pw) in enumerate(storeys):
    rr = ST_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, pw, BLUE, YEL, RATE)
ST_E = ST_S + len(storeys) - 1

# --- Pitch ---
r = ST_E + 2
sec_hdr(cp, r, "PITCH ADDERS", 2)
r += 1
sc(cp, r, 1, "Slope", BLK_B, LGRAY); sc(cp, r, 2, "Per-Watt ($)", BLK_B, LGRAY)
pitches = [("Standard (< 34\u00b0)", 0), ("Steep (34\u00b0-44\u00b0)", 0.35), ("Very Steep (> 44\u00b0)", 0.50)]
PI_S = r + 1
for i, (n, pw) in enumerate(pitches):
    rr = PI_S + i
    sc(cp, rr, 1, n, BLUE)
    sc(cp, rr, 2, pw, BLUE, YEL, RATE)
PI_E = PI_S + len(pitches) - 1

# --- Cost Scheme Rates ---
r = PI_E + 2
sec_hdr(cp, r, "COST SCHEME RATES", 3)
r += 1
sc(cp, r, 1, "Item", BLK_B, LGRAY); sc(cp, r, 2, "Rate", BLK_B, LGRAY); sc(cp, r, 3, "Unit", BLK_B, LGRAY)
rates = [
    ("Racking", 0.15, "$/W"), ("BOS", 0.15, "$/W"), ("Labour", 0.55, "$/W"),
    ("Lead Gen (fixed)", 300, "$/system"), ("Lead Gen (per-watt)", 0.10, "$/W"),
    ("Salary", 100, "$/system"), ("Commission", 0.05, "% of COGS+Labour"),
    ("Pre-sale", 0.01, "$/W"), ("PM", 1000, "$/system"),
    ("Design", 350, "$/system"), ("Permit", 500, "$/system"), ("Battery Misc", 200, "$/system"),
]
RT_S = r + 1
for i, (n, v, u) in enumerate(rates):
    rr = RT_S + i
    sc(cp, rr, 1, n, BLK); sc(cp, rr, 2, v, BLUE, YEL, RATE if v < 10 else CUR); sc(cp, rr, 3, u, BLK)
RT_E = RT_S + len(rates) - 1
# Named rate row offsets
RR = {name: RT_S + i for i, (name, _, _) in enumerate(rates)}

# --- PE Lease ---
r = RT_E + 2
sec_hdr(cp, r, "PE LEASE PARAMETERS", 3)
r += 1
sc(cp, r, 1, "Parameter", BLK_B, LGRAY); sc(cp, r, 2, "Value", BLK_B, LGRAY); sc(cp, r, 3, "Notes", BLK_B, LGRAY)
pe_params = [
    ("Baseline Factor", 1.4285714, "10/7 per PE Pricing Policy v3"),
    ("DC Bonus", 0.1098901, "When domestic content qualifies"),
    ("No Bonus Penalty", -0.0952381, "No DC and no Energy Community"),
    ("Customer Discount", 0.30, "PE customer pays 70%"),
]
PE_S = r + 1
for i, (n, v, note) in enumerate(pe_params):
    rr = PE_S + i
    sc(cp, rr, 1, n, BLK); sc(cp, rr, 2, v, BLUE, YEL, FAC if abs(v) > 0.5 else RATE); sc(cp, rr, 3, note, BLK)
PE_BASE = PE_S
PE_DCB = PE_S + 1
PE_NOB = PE_S + 2
PE_DISC = PE_S + 3

C = "'Costs & Prices'!"

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2: Calculator
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Calculator")
ws.sheet_properties.tabColor = "FF6600"
wb.active = wb.sheetnames.index("Calculator")

for c, w in enumerate([28, 28, 10, 16, 16], 1):
    ws.column_dimensions[get_column_letter(c)].width = w

# Title
r = 1
for c in range(1, 6):
    sc(ws, r, c, "PB Pricing Calculator" if c == 1 else None, WHT_B, HDR)
ws.merge_cells("A1:E1")
ws["A1"].alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════════════════════════════════════════════
# INPUT SECTION
# ═══════════════════════════════════════════════════════════════════════════════
r = 3
sec_hdr(ws, r, "EQUIPMENT")
r += 1
sc(ws, r, 1, "", BLK_B, LGRAY); sc(ws, r, 2, "Selection", BLK_B, LGRAY); sc(ws, r, 3, "Qty", BLK_B, LGRAY)
sc(ws, r, 4, "Cost/Unit", BLK_B, LGRAY); sc(ws, r, 5, "Total", BLK_B, LGRAY)

# Module 1
r += 1; M1R = r
sc(ws, r, 1, "Module", BLK_B, LBLU)
sc(ws, r, 2, modules[0][0], BLUE_B, YEL)
sc(ws, r, 3, 0, BLUE, YEL, NUM)
sc(ws, r, 4, f"=VLOOKUP(B{r},{C}A{MOD_S}:B{MOD_E},2,FALSE)", BLK, LBLU, CUR)
sc(ws, r, 5, f"=C{r}*D{r}", BLK, LBLU, CUR)

# Module 2
r += 1; M2R = r
sc(ws, r, 1, "Module 2 (optional)", BLK, LBLU)
sc(ws, r, 2, "(none)", BLUE_B, YEL)
sc(ws, r, 3, 0, BLUE, YEL, NUM)
sc(ws, r, 4, f'=IF(B{r}="(none)",0,VLOOKUP(B{r},{C}A{MOD_S}:B{MOD_E},2,FALSE))', BLK, LBLU, CUR)
sc(ws, r, 5, f"=C{r}*D{r}", BLK, LBLU, CUR)

# Inverter 1
r += 1; I1R = r
sc(ws, r, 1, "Inverter", BLK_B)
sc(ws, r, 2, inverters[0][0], BLUE_B, YEL)
sc(ws, r, 3, 0, BLUE, YEL, NUM)
sc(ws, r, 4, f"=VLOOKUP(B{r},{C}A{INV_S}:B{INV_E},2,FALSE)", BLK, None, CUR)
sc(ws, r, 5, f"=C{r}*D{r}", BLK, None, CUR)

# Inverter 2
r += 1; I2R = r
sc(ws, r, 1, "Inverter 2 (optional)", BLK)
sc(ws, r, 2, "(none)", BLUE_B, YEL)
sc(ws, r, 3, 0, BLUE, YEL, NUM)
sc(ws, r, 4, f'=IF(B{r}="(none)",0,VLOOKUP(B{r},{C}A{INV_S}:B{INV_E},2,FALSE))', BLK, None, CUR)
sc(ws, r, 5, f"=C{r}*D{r}", BLK, None, CUR)

# Battery 1
r += 1; B1R = r
sc(ws, r, 1, "Battery", BLK_B, LGRN)
sc(ws, r, 2, "(none)", BLUE_B, YEL)
sc(ws, r, 3, 0, BLUE, YEL, NUM)
sc(ws, r, 4, f'=IF(B{r}="(none)",0,VLOOKUP(B{r},{C}A{BAT_S}:B{BAT_E},2,FALSE))', BLK, LGRN, CUR)
sc(ws, r, 5, f"=C{r}*D{r}", BLK, LGRN, CUR)

# Battery 2
r += 1; B2R = r
sc(ws, r, 1, "Battery 2 (optional)", BLK, LGRN)
sc(ws, r, 2, "(none)", BLUE_B, YEL)
sc(ws, r, 3, 0, BLUE, YEL, NUM)
sc(ws, r, 4, f'=IF(B{r}="(none)",0,VLOOKUP(B{r},{C}A{BAT_S}:B{BAT_E},2,FALSE))', BLK, LGRN, CUR)
sc(ws, r, 5, f"=C{r}*D{r}", BLK, LGRN, CUR)

# EV Charger
r += 1; EVR = r
sc(ws, r, 1, "EV Charger", BLK)
sc(ws, r, 2, "No", BLUE_B, YEL)
sc(ws, r, 3, 1, BLUE, YEL, NUM)
sc(ws, r, 4, f"={C}B{OTH_S}", BLK, None, CUR)
sc(ws, r, 5, f'=IF(B{r}="Yes",C{r}*D{r},0)', BLK, None, CUR)

# Data validations for equipment
mod_list = ",".join([m[0] for m in modules])
mod_dv = DataValidation(type="list", formula1=f"{C}$A${MOD_S}:$A${MOD_E}", allow_blank=False, showDropDown=False)
mod2_list = "(none)," + ",".join([m[0] for m in modules])
mod2_dv = DataValidation(type="list", formula1=f'"(none),{mod_list}"', allow_blank=False, showDropDown=False)
inv_dv = DataValidation(type="list", formula1=f"{C}$A${INV_S}:$A${INV_E}", allow_blank=False, showDropDown=False)
inv_list = ",".join([i[0] for i in inverters])
inv2_dv = DataValidation(type="list", formula1=f'"(none),{inv_list}"', allow_blank=False, showDropDown=False)
bat_list = ",".join([b[0] for b in batteries])
bat_dv = DataValidation(type="list", formula1=f'"(none),{bat_list}"', allow_blank=False, showDropDown=False)
yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False, showDropDown=False)

ws.add_data_validation(mod_dv); mod_dv.add(ws.cell(M1R, 2))
ws.add_data_validation(mod2_dv); mod2_dv.add(ws.cell(M2R, 2))
ws.add_data_validation(inv_dv); inv_dv.add(ws.cell(I1R, 2))
ws.add_data_validation(inv2_dv); inv2_dv.add(ws.cell(I2R, 2))
ws.add_data_validation(bat_dv); bat_dv.add(ws.cell(B1R, 2)); bat_dv.add(ws.cell(B2R, 2))
ws.add_data_validation(yn_dv); yn_dv.add(ws.cell(EVR, 2))

# ─── Configuration ────────────────────────────────────────────────────────────
r += 2
sec_hdr(ws, r, "CONFIGURATION")
r += 1
sc(ws, r, 1, "Pricing Scheme", BLK_B); sc(ws, r, 2, "Base (Colorado)", BLUE_B, YEL)
sc(ws, r, 3, "Markup:", BLK); sc(ws, r, 4, f"=VLOOKUP(B{r},{C}A{SCH_S}:B{SCH_E},2,FALSE)", BLK_B, None, PCT)
SCH_R = r
sch_dv = DataValidation(type="list", formula1=f"{C}$A${SCH_S}:$A${SCH_E}", allow_blank=False, showDropDown=False)
ws.add_data_validation(sch_dv); sch_dv.add(ws.cell(r, 2))

r += 1
sc(ws, r, 1, "Roof Type", BLK_B); sc(ws, r, 2, "Comp/Asphalt Shingle", BLUE_B, YEL)
sc(ws, r, 3, "Fixed:", BLK)
sc(ws, r, 4, f"=VLOOKUP(B{r},{C}A{RF_S}:C{RF_E},2,FALSE)", BLK, None, CUR)
sc(ws, r, 5, f"=VLOOKUP(B{r},{C}A{RF_S}:C{RF_E},3,FALSE)", BLK, None, RATE)
RF_R = r
rf_dv = DataValidation(type="list", formula1=f"{C}$A${RF_S}:$A${RF_E}", allow_blank=False, showDropDown=False)
ws.add_data_validation(rf_dv); rf_dv.add(ws.cell(r, 2))

r += 1
sc(ws, r, 1, "Storey", BLK_B); sc(ws, r, 2, "1 Story", BLUE_B, YEL)
sc(ws, r, 3, "$/W:", BLK); sc(ws, r, 4, f"=VLOOKUP(B{r},{C}A{ST_S}:B{ST_E},2,FALSE)", BLK, None, RATE)
ST_R = r
st_dv = DataValidation(type="list", formula1=f"{C}$A${ST_S}:$A${ST_E}", allow_blank=False, showDropDown=False)
ws.add_data_validation(st_dv); st_dv.add(ws.cell(r, 2))

r += 1
sc(ws, r, 1, "Roof Pitch", BLK_B); sc(ws, r, 2, "Standard (< 34\u00b0)", BLUE_B, YEL)
sc(ws, r, 3, "$/W:", BLK); sc(ws, r, 4, f"=VLOOKUP(B{r},{C}A{PI_S}:B{PI_E},2,FALSE)", BLK, None, RATE)
PI_R = r
pi_dv = DataValidation(type="list", formula1=f"{C}$A${PI_S}:$A${PI_E}", allow_blank=False, showDropDown=False)
ws.add_data_validation(pi_dv); pi_dv.add(ws.cell(r, 2))

r += 1
sc(ws, r, 1, "PE Deal?", BLK_B); sc(ws, r, 2, "No", BLUE_B, YEL)
PE_R = r
yn_dv.add(ws.cell(r, 2))

r += 1
sc(ws, r, 1, "Energy Community?", BLK_B); sc(ws, r, 2, "No", BLUE_B, YEL)
EC_R = r
yn_dv.add(ws.cell(r, 2))

r += 1
sc(ws, r, 1, "Custom Adder ($)", BLK_B); sc(ws, r, 2, 0, BLUE, YEL, CUR)
sc(ws, r, 3, "e.g. -1000 promo", HINT)
CA_R = r

# ═══════════════════════════════════════════════════════════════════════════════
# COMPUTED HELPERS (column F — watts & labour lookups)
# ═══════════════════════════════════════════════════════════════════════════════
ws.column_dimensions["F"].width = 2
ws.column_dimensions["F"].hidden = True

# F column: watts for module selections
sc(ws, M1R, 6, f'=VLOOKUP(B{M1R},{C}A{MOD_S}:C{MOD_E},3,FALSE)', BLK)
sc(ws, M2R, 6, f'=IF(B{M2R}="(none)",0,VLOOKUP(B{M2R},{C}A{MOD_S}:C{MOD_E},3,FALSE))', BLK)
# F column: battery labour
sc(ws, B1R, 6, f'=IF(B{B1R}="(none)",0,VLOOKUP(B{B1R},{C}A{BAT_S}:C{BAT_E},3,FALSE))', BLK)
sc(ws, B2R, 6, f'=IF(B{B2R}="(none)",0,VLOOKUP(B{B2R},{C}A{BAT_S}:C{BAT_E},3,FALSE))', BLK)

# ═══════════════════════════════════════════════════════════════════════════════
# SYSTEM SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
r = CA_R + 2
sec_hdr(ws, r, "SYSTEM SUMMARY")
r += 1; PANELS_R = r
sc(ws, r, 1, "Total Panels", BLK_B)
sc(ws, r, 2, f"=C{M1R}+C{M2R}", BLK, None, NUM)
r += 1; WATTS_R = r
sc(ws, r, 1, "Total Watts", BLK_B)
sc(ws, r, 2, f"=C{M1R}*F{M1R}+C{M2R}*F{M2R}", BLK, None, NUM)
sc(ws, r, 3, f"=B{r}/1000", BLK, None, "0.00")
sc(ws, r, 4, "kW DC", HINT)
W = f"B{WATTS_R}"
r += 1; BATCNT_R = r
sc(ws, r, 1, "Total Batteries", BLK_B)
sc(ws, r, 2, f"=C{B1R}+C{B2R}", BLK, None, NUM)
BC = f"B{BATCNT_R}"
r += 1
sc(ws, r, 1, "Battery Only?", BLK_B)
sc(ws, r, 2, f'=IF(AND(B{PANELS_R}=0,{BC}>0),"Yes","No")', BLK)
BONLY_R = r

# ═══════════════════════════════════════════════════════════════════════════════
# COST BREAKDOWN
# ═══════════════════════════════════════════════════════════════════════════════
r += 2
sec_hdr(ws, r, "COST BREAKDOWN")
r += 1; BD_S = r
sc(ws, r, 1, "Module Cost", BLK); sc(ws, r, 5, f"=E{M1R}+E{M2R}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Inverter Cost", BLK); sc(ws, r, 5, f"=E{I1R}+E{I2R}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Battery Cost", BLK); sc(ws, r, 5, f"=E{B1R}+E{B2R}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Other Equipment", BLK); sc(ws, r, 5, f"=E{EVR}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Battery Misc", BLK)
sc(ws, r, 5, f'=IF(OR(B{BONLY_R}="Yes",{BC}>=2),IF({BC}>0,{C}B{RR["Battery Misc"]},0),0)', BLK, None, CUR)
r += 1
sc(ws, r, 1, "Racking", BLK); sc(ws, r, 5, f"={W}*{C}B{RR['Racking']}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "BOS", BLK); sc(ws, r, 5, f"={W}*{C}B{RR['BOS']}", BLK, None, CUR)
r += 1; COGS_R = r
sc(ws, r, 1, "COGS", BLK_B); sc(ws, r, 5, f"=SUM(E{BD_S}:E{r-1})", BLK_B, None, CUR)

r += 2; EX_S = r
sc(ws, r, 1, "Roof Adder", BLK); sc(ws, r, 5, f"=D{RF_R}+{W}*E{RF_R}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Storey Adder", BLK); sc(ws, r, 5, f"={W}*D{ST_R}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Pitch Adder", BLK); sc(ws, r, 5, f"={W}*D{PI_R}", BLK, None, CUR)
r += 1; EX_T = r
sc(ws, r, 1, "Extra Costs", BLK_B); sc(ws, r, 5, f"=SUM(E{EX_S}:E{r-1})", BLK_B, None, CUR)

r += 2; LB_S = r
sc(ws, r, 1, "General Labour", BLK); sc(ws, r, 5, f"={W}*{C}B{RR['Labour']}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Battery Labour", BLK); sc(ws, r, 5, f"=C{B1R}*F{B1R}+C{B2R}*F{B2R}", BLK, None, CUR)
r += 1; LB_T = r
sc(ws, r, 1, "Total Labour", BLK_B); sc(ws, r, 5, f"=SUM(E{LB_S}:E{r-1})", BLK_B, None, CUR)

r += 2; AQ_S = r
sc(ws, r, 1, "Lead Gen", BLK); sc(ws, r, 5, f"={C}B{RR['Lead Gen (fixed)']}+{W}*{C}B{RR['Lead Gen (per-watt)']}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Salary", BLK); sc(ws, r, 5, f"={C}B{RR['Salary']}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Commission (5%)", BLK); sc(ws, r, 5, f"=(E{COGS_R}+E{LB_T})*{C}B{RR['Commission']}", BLK, None, CUR)
r += 1
sc(ws, r, 1, "Pre-sale", BLK); sc(ws, r, 5, f"={W}*{C}B{RR['Pre-sale']}", BLK, None, CUR)
r += 1; AQ_T = r
sc(ws, r, 1, "Total Acquisition", BLK_B); sc(ws, r, 5, f"=SUM(E{AQ_S}:E{r-1})", BLK_B, None, CUR)

r += 2; FF_R = r
sc(ws, r, 1, "Fulfillment (PM+Design+Permit)", BLK_B)
sc(ws, r, 5, f"={C}B{RR['PM']}+{C}B{RR['Design']}+{C}B{RR['Permit']}", BLK_B, None, CUR)

# ─── Totals ───────────────────────────────────────────────────────────────────
r += 2; TC_R = r
for c in range(1, 6): sc(ws, r, c, None, BLK_B, LORG)
sc(ws, r, 1, "TOTAL COSTS", BLK_B, LORG)
sc(ws, r, 5, f"=E{COGS_R}+E{EX_T}+E{LB_T}+E{AQ_T}+E{FF_R}", BLK_B, LORG, CUR)

r += 1; BP_R = r
for c in range(1, 6): sc(ws, r, c, None, BLK_B, LORG)
sc(ws, r, 1, "BASE PRICE (Costs x (1+Markup))", BLK_B, LORG)
sc(ws, r, 5, f"=E{TC_R}*(1+D{SCH_R})", BLK_B, LORG, CUR)

r += 1; AF_R = r
sc(ws, r, 1, "AFTER ADDERS", BLK_B)
sc(ws, r, 5, f"=E{BP_R}+B{CA_R}", BLK_B, None, CUR)

# ═══════════════════════════════════════════════════════════════════════════════
# PE LEASE & FINAL PRICING
# ═══════════════════════════════════════════════════════════════════════════════
r += 2
sec_hdr(ws, r, "PE LEASE & FINAL PRICING")

r += 1; ST_TYPE = r
sc(ws, r, 1, "System Type", BLK_B)
sc(ws, r, 2, f'=IF(AND(B{PANELS_R}>0,{BC}>0),"solar+battery",IF(B{PANELS_R}>0,"solar",IF({BC}>0,"battery","none")))', BLK)

r += 1
sc(ws, r, 1, "Solar DC Qualified?", BLK_B)
# Auto-detect: all selected modules must have DC Qual?=Yes on Costs & Prices
sc(ws, r, 2,
    f'=IF(B{PANELS_R}=0,"No",'
    f'IF(AND('
    f'IF(C{M1R}>0,VLOOKUP(B{M1R},{C}A{MOD_S}:D{MOD_E},4,FALSE)="Yes",TRUE),'
    f'IF(C{M2R}>0,VLOOKUP(B{M2R},{C}A{MOD_S}:D{MOD_E},4,FALSE)="Yes",TRUE)'
    f'),"Yes","No"))',
    BLK)
SDC_R = r

r += 1
sc(ws, r, 1, "Battery DC Qualified?", BLK_B)
# Auto-detect: all selected batteries must have DC Qual?=Yes on Costs & Prices
sc(ws, r, 2,
    f'=IF({BC}=0,"No",'
    f'IF(AND('
    f'IF(C{B1R}>0,VLOOKUP(B{B1R},{C}A{BAT_S}:D{BAT_E},4,FALSE)="Yes",TRUE),'
    f'IF(C{B2R}>0,VLOOKUP(B{B2R},{C}A{BAT_S}:D{BAT_E},4,FALSE)="Yes",TRUE)'
    f'),"Yes","No"))',
    BLK)
BDC_R = r

r += 1; ADJ_R = r
sc(ws, r, 1, "Lease Factor Adjustment", BLK_B)
sc(ws, r, 2,
    f'=IF(B{PE_R}<>"Yes",0,'
    f'IF(B{ST_TYPE}="solar+battery",'
        f'IF(AND(B{SDC_R}="Yes",B{BDC_R}="Yes"),{C}B{PE_DCB},'
        f'IF(OR(B{SDC_R}="Yes",B{BDC_R}="Yes",B{EC_R}="Yes"),0,{C}B{PE_NOB})),'
    f'IF(B{ST_TYPE}="solar",'
        f'IF(B{SDC_R}="Yes",{C}B{PE_DCB},'
        f'IF(B{EC_R}="Yes",0,{C}B{PE_NOB})),'
    f'IF(B{ST_TYPE}="battery",'
        f'IF(B{BDC_R}="Yes",{C}B{PE_DCB},'
        f'IF(B{EC_R}="Yes",0,{C}B{PE_NOB})),0))))',
    BLK, None, FAC)

r += 1; FAC_R = r
for c in range(1, 6): sc(ws, r, c, None, BLK_B, LORG)
sc(ws, r, 1, "Lease Factor", BLK_B, LORG)
sc(ws, r, 2, f"={C}B{PE_BASE}+B{ADJ_R}", BLK_B, LORG, FAC)

r += 2; EPC_R = r
for c in range(1, 6): sc(ws, r, c, None, BLK_B, LORG)
sc(ws, r, 1, "EPC Price (HubSpot Amount)", BLK_B, LORG)
sc(ws, r, 5, f"=E{AF_R}", BLK_B, LORG, CUR2)

r += 1; CUST_R = r
for c in range(1, 6): sc(ws, r, c, None, BLK_B, LGRN)
sc(ws, r, 1, "CUSTOMER PAYS", BLK_B, LGRN)
sc(ws, r, 5, f'=IF(B{PE_R}="Yes",E{EPC_R}*(1-{C}B{PE_DISC}),E{EPC_R})', BLK_B, LGRN, CUR2)

r += 1; PET_R = r
sc(ws, r, 1, "PE Payment to PB", BLK_B)
sc(ws, r, 5, f'=IF(B{PE_R}="Yes",E{EPC_R}-(E{EPC_R}/B{FAC_R}),0)', BLK_B, None, CUR2)

r += 1
sc(ws, r, 1, "  PE @ Inspection Complete (2/3)", BLK)
sc(ws, r, 5, f"=E{PET_R}*(2/3)", BLK, None, CUR2)

r += 1
sc(ws, r, 1, "  PE @ Project Complete (1/3)", BLK)
sc(ws, r, 5, f"=E{PET_R}*(1/3)", BLK, None, CUR2)

r += 1; REV_R = r
for c in range(1, 6): sc(ws, r, c, None, BLK_B, LORG)
sc(ws, r, 1, "TOTAL PB REVENUE", BLK_B, LORG)
sc(ws, r, 5, f'=IF(B{PE_R}="Yes",E{CUST_R}+E{PET_R},E{CUST_R})', BLK_B, LORG, CUR2)

# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
OUTPUT = "/Users/zach/Downloads/Dev Projects/PB-Operations-Suite/PB-Pricing-Calculator.xlsx"
wb.save(OUTPUT)
print(f"Saved to {OUTPUT}")
